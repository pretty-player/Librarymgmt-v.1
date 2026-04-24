import tkinter as tk
from tkinter import messagebox, ttk, filedialog, simpledialog
import sqlite3
import shutil
import os
from datetime import datetime

# Dependency Check for Excel Support
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

class LibrarySystem:
    def __init__(self, root):
        self.root = root
        self.root.title(" Library System v1")
        self.root.geometry("1300x750")

        # Configuration
        self.ADMIN_PASSWORD = "admin123" 
        self.pdf_folder = "library_pdfs"
        self.log_file = "system_logs.txt"
        
        # Directory Initialization
        if not os.path.exists(self.pdf_folder):
            os.makedirs(self.pdf_folder)

        # Database Initialization
        self.conn = sqlite3.connect("library.db")
        self.cursor = self.conn.cursor()
        self.setup_db()

        # Tkinter Variables
        self.student_name = tk.StringVar()
        self.search_query = tk.StringVar()
        
        self.create_main_widgets()
        self.refresh_main_view()
        self.write_log("System Session Started")

    def setup_db(self):
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS books (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT,
                author TEXT,
                borrower TEXT DEFAULT 'Available',
                pdf_path TEXT DEFAULT '',
                date_added TEXT,
                issue_date TEXT DEFAULT 'N/A'
            )
        """)
        self.conn.commit()

    def write_log(self, action):
        """Internal method to record every action with a precise timestamp."""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(self.log_file, "a", encoding="utf-8") as f:
            f.write(f"[{now}] {action}\n")

    def create_main_widgets(self):
        # --- Top Navigation Bar ---
        nav_frame = tk.Frame(self.root, pady=10, bg="#2c3e50")
        nav_frame.pack(fill="x")
        
        btn_style = {"font": ('Arial', 9, 'bold'), "padx": 10, "fg": "white"}
        
        tk.Button(nav_frame, text="➕ Add Book", command=self.open_book_window, bg="#3498db", **btn_style).pack(side="left", padx=10)
        
        if EXCEL_SUPPORT:
            tk.Button(nav_frame, text="📥 Bulk Upload", command=self.bulk_upload_excel, bg="#e67e22", **btn_style).pack(side="left", padx=5)
            tk.Button(nav_frame, text="📊 Export Excel", command=self.export_to_excel, bg="#27ae60", **btn_style).pack(side="left", padx=5)

        # Admin Controls
        tk.Button(nav_frame, text="📜 View Logs", command=self.open_log_viewer, bg="#9b59b6", **btn_style).pack(side="left", padx=5)
        tk.Button(nav_frame, text="🗑️ Factory Reset", command=self.admin_clear_all, bg="#c0392b", **btn_style).pack(side="left", padx=20)

        # Quick Search UI
        search_frame = tk.Frame(nav_frame, bg="#2c3e50")
        search_frame.pack(side="right", padx=20)
        search_entry = tk.Entry(search_frame, textvariable=self.search_query, width=25)
        search_entry.pack(side="left", padx=5)
        search_entry.bind('<Return>', lambda e: self.search_books())
        tk.Button(search_frame, text="🔍", command=self.search_books, bg="#bdc3c7").pack(side="left")

        # --- Dashboard Operations ---
        ops_frame = tk.LabelFrame(self.root, text="Student Circulation & Digital Assets", padx=10, pady=10)
        ops_frame.pack(fill="x", padx=20, pady=10)

        tk.Label(ops_frame, text="Student:").grid(row=0, column=0)
        tk.Entry(ops_frame, textvariable=self.student_name, width=20).grid(row=0, column=1, padx=5)
        
        tk.Button(ops_frame, text="Issue Book", command=self.issue_book, width=10).grid(row=0, column=2, padx=2)
        tk.Button(ops_frame, text="Return Book", command=self.return_book, width=10).grid(row=0, column=3, padx=2)
        tk.Button(ops_frame, text="📎 Link PDF", command=self.link_pdf_to_selected, bg="#34495e", fg="white", width=12).grid(row=0, column=4, padx=15)
        tk.Button(ops_frame, text="📖 View PDF", command=self.open_pdf, bg="#f1c40f", width=12).grid(row=0, column=5, padx=5)

        # --- Main Data Table (Treeview) ---
        cols = ("ID", "Title", "Author", "Status", "Added On", "Issue Date", "PDF", "Action")
        self.tree = ttk.Treeview(self.root, columns=cols, show="headings")
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=140, anchor="center")
        self.tree.column("ID", width=40); self.tree.column("PDF", width=50); self.tree.column("Action", width=110)
        self.tree.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Interaction Binding
        self.tree.bind("<Double-1>", lambda e: self.delete_book_item())

    # --- THE LOG VIEWER ---
    def open_log_viewer(self):
        """Displays historical system actions in a classical console-style window."""
        pwd = simpledialog.askstring("Admin Required", "Enter Password to Access Logs:", show='*')
        if pwd != self.ADMIN_PASSWORD:
            messagebox.showerror("Error", "Unauthorized Access Attempt Blocked.")
            self.write_log("FAILED ACCESS: Unauthorized attempt to view logs.")
            return

        log_win = tk.Toplevel(self.root)
        log_win.title("Classical System Log Viewer")
        log_win.geometry("700x550")

        # Terminal-style Text Area
        text_area = tk.Text(log_win, padx=15, pady=15, bg="#1e1e1e", fg="#00ff41", font=("Courier", 10))
        text_area.pack(fill="both", expand=True)

        if os.path.exists(self.log_file):
            with open(self.log_file, "r", encoding="utf-8") as f:
                text_area.insert("1.0", f.read())
        
        text_area.config(state="disabled") # Prevent accidental editing of logs

        btn_frame = tk.Frame(log_win, pady=10)
        btn_frame.pack(fill="x")

        def clear_logs():
            if messagebox.askyesno("Confirm Wipe", "Permanently delete all log history?"):
                open(self.log_file, "w").close()
                text_area.config(state="normal")
                text_area.delete("1.0", "end")
                text_area.config(state="disabled")
                self.write_log("LOG HISTORY CLEARED: Action performed by Admin.")

        tk.Button(btn_frame, text="Clear Log File", command=clear_logs, bg="#e74c3c", fg="white").pack(side="right", padx=20)
        tk.Button(btn_frame, text="Close Viewer", command=log_win.destroy).pack(side="right")

    # --- BUSINESS LOGIC WITH LOGGING ---

    def issue_book(self):
        sel = self.tree.selection()
        if sel and self.student_name.get().strip():
            bid = self.tree.item(sel[0])['values'][0]
            title = self.tree.item(sel[0])['values'][1]
            student = self.student_name.get().strip()
            
            self.cursor.execute("UPDATE books SET borrower=?, issue_date=? WHERE id=?", 
                               (student, datetime.now().strftime("%Y-%m-%d %H:%M"), bid))
            self.conn.commit()
            self.write_log(f"ISSUE: '{title}' (ID:{bid}) assigned to {student}")
            self.refresh_main_view()
            self.student_name.set("")
        else:
            messagebox.showwarning("Incomplete", "Please select a book and enter student name.")

    def return_book(self):
        sel = self.tree.selection()
        if sel:
            bid = self.tree.item(sel[0])['values'][0]
            title = self.tree.item(sel[0])['values'][1]
            self.cursor.execute("UPDATE books SET borrower='Available', issue_date='N/A' WHERE id=?", (bid,))
            self.conn.commit()
            self.write_log(f"RETURN: '{title}' (ID:{bid}) returned to inventory")
            self.refresh_main_view()

    def delete_book_item(self):
        sel = self.tree.selection()
        if sel:
            data = self.tree.item(sel[0])['values']
            if messagebox.askyesno("Delete", f"Are you sure you want to delete '{data[1]}'?"):
                self.cursor.execute("DELETE FROM books WHERE id=?", (data[0],))
                self.conn.commit()
                self.write_log(f"DELETE: Record ID {data[0]} ('{data[1]}') removed.")
                self.refresh_main_view()

    def bulk_upload_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not path: return
        try:
            wb = load_workbook(path); sheet = wb.active
            count = 0
            now = datetime.now().strftime("%Y-%m-%d %H:%M")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]: # Title check
                    self.cursor.execute("INSERT INTO books (title, author, date_added) VALUES (?, ?, ?)", 
                                       (str(row[0]), str(row[1]), now))
                    count += 1
            self.conn.commit()
            self.write_log(f"BULK UPLOAD: {count} books imported via Excel.")
            self.refresh_main_view()
            messagebox.showinfo("Success", f"Imported {count} records.")
        except Exception as e:
            messagebox.showerror("Excel Error", str(e))

    def export_to_excel(self):
        if not EXCEL_SUPPORT: return
        wb = Workbook(); ws = wb.active
        ws.append(["ID", "Title", "Author", "Status", "Date Added", "Last Issue", "PDF Linked"])
        self.cursor.execute("SELECT id, title, author, borrower, date_added, issue_date, pdf_path FROM books")
        for r in self.cursor.fetchall():
            pdf_status = "Yes" if r[6] else "No"
            ws.append(list(r[:6]) + [pdf_status])
        
        fname = f"Library_Inventory_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        wb.save(fname)
        self.write_log(f"EXPORT: Full inventory exported to {fname}")
        messagebox.showinfo("Export Complete", f"Data saved to {fname}")

    def refresh_main_view(self, data=None):
        for item in self.tree.get_children(): self.tree.delete(item)
        if data is None:
            self.cursor.execute("SELECT id, title, author, borrower, date_added, issue_date, pdf_path FROM books")
            data = self.cursor.fetchall()
        for row in data:
            pdf_stat = "✅" if (row[6] and os.path.exists(row[6])) else "❌"
            self.tree.insert("", "end", values=(row[0], row[1], row[2], row[3], row[4], row[5], pdf_stat, "🗑️ Double-Click"))

    def open_book_window(self):
        win = tk.Toplevel(self.root); win.title("Manual Entry"); win.geometry("300x350")
        t_var, a_var = tk.StringVar(), tk.StringVar(); self.temp_p = ""
        def select_pdf(): self.temp_p = filedialog.askopenfilename(filetypes=[("PDF", "*.pdf")])
        def save():
            if t_var.get().strip():
                dest = ""
                if self.temp_p:
                    dest = os.path.join(self.pdf_folder, f"lib_{datetime.now().strftime('%M%S')}.pdf")
                    shutil.copy(self.temp_p, dest)
                self.cursor.execute("INSERT INTO books (title, author, pdf_path, date_added) VALUES (?,?,?,?)", 
                                   (t_var.get(), a_var.get(), dest, datetime.now().strftime("%Y-%m-%d %H:%M")))
                self.conn.commit()
                self.write_log(f"ADD: Manual entry created for '{t_var.get()}'")
                self.refresh_main_view(); win.destroy()
        tk.Label(win, text="Title:").pack(pady=5); tk.Entry(win, textvariable=t_var).pack()
        tk.Label(win, text="Author:").pack(pady=5); tk.Entry(win, textvariable=a_var).pack()
        tk.Button(win, text="📎 Select PDF", command=select_pdf).pack(pady=10)
        tk.Button(win, text="Save Record", command=save, bg="#27ae60", fg="white").pack(pady=20)

    def link_pdf_to_selected(self):
        sel = self.tree.selection()
        if not sel: return
        bid = self.tree.item(sel[0])['values'][0]
        title = self.tree.item(sel[0])['values'][1]
        path = filedialog.askopenfilename(filetypes=[("PDF", "*.pdf")])
        if path:
            dest = os.path.join(self.pdf_folder, f"link_{bid}.pdf")
            shutil.copy(path, dest)
            self.cursor.execute("UPDATE books SET pdf_path=? WHERE id=?", (dest, bid))
            self.conn.commit()
            self.write_log(f"LINK: Attached PDF to '{title}' (ID:{bid})")
            self.refresh_main_view()

    def search_books(self):
        q = self.search_query.get().strip()
        if not q: self.refresh_main_view(); return
        term = f"%{q}%"
        self.cursor.execute("SELECT id, title, author, borrower, date_added, issue_date, pdf_path FROM books WHERE title LIKE ? OR author LIKE ? OR borrower LIKE ?", (term, term, term))
        self.refresh_main_view(data=self.cursor.fetchall())

    def open_pdf(self):
        sel = self.tree.selection()
        if not sel: return
        self.cursor.execute("SELECT pdf_path, title FROM books WHERE id=?", (self.tree.item(sel[0])['values'][0],))
        res = self.cursor.fetchone()
        if res[0] and os.path.exists(res[0]):
            os.startfile(res[0])
            self.write_log(f"VIEW: PDF for '{res[1]}' opened.")
        else:
            messagebox.showinfo("Digital Asset", "No PDF file is currently linked to this record.")

    def admin_clear_all(self):
        if simpledialog.askstring("Admin Security", "Enter Password to CLEAR ALL:", show='*') == self.ADMIN_PASSWORD:
            if messagebox.askyesno("Critical Confirm", "This will WIPE all records and DELETE all PDFs. Continue?"):
                self.cursor.execute("DELETE FROM books"); self.conn.commit()
                for f in os.listdir(self.pdf_folder): os.remove(os.path.join(self.pdf_folder, f))
                self.write_log("FACTORY RESET: ALL DATA WIPED BY ADMIN.")
                self.refresh_main_view()
        else:
            messagebox.showerror("Error", "Incorrect Password.")

if __name__ == "__main__":
    root = tk.Tk(); app = LibrarySystem(root); root.mainloop()